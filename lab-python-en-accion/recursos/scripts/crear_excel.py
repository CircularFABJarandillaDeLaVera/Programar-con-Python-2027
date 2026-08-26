from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


def crear_excel_ventas():
    base_dir = Path(__file__).resolve().parents[1]
    salida = base_dir / "ventas_lab.xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Ventas"

    ws.append(["Producto", "Unidades", "Precio", "Importe"])
    ws.append(["Kit robotica", 2, 49.90, "=B2*C2"])
    ws.append(["Sensor distancia", 5, 12.50, "=B3*C3"])
    ws.append(["Cable USB", 10, 3.20, "=B4*C4"])
    ws.append(["Placa control", 3, 24.75, "=B5*C5"])
    ws["C7"] = "Total"
    ws["D7"] = "=SUM(D2:D5)"

    cabecera = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
    destacado = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    alerta = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")

    for celda in ws[1]:
        celda.font = Font(bold=True, color="FFFFFF")
        celda.fill = cabecera

    for fila in ws.iter_rows(min_row=2, max_row=5, min_col=1, max_col=4):
        unidades = fila[1].value
        for celda in fila:
            celda.fill = alerta if unidades >= 10 else destacado

    ws["C7"].font = Font(bold=True)
    ws["D7"].font = Font(bold=True)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14

    wb.save(salida)
    print(f"Excel generado: {salida}")


if __name__ == "__main__":
    crear_excel_ventas()
